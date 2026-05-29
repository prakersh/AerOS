"""Spreadsheet extractor — handles xlsx, xls, csv, tsv."""

import csv
import io
from typing import Any


async def extract_spreadsheet(file_path: str, **kwargs: Any) -> str:
    if file_path.endswith((".csv", ".tsv")):
        return await _extract_csv(file_path)
    return await _extract_excel(file_path)


async def _extract_excel(file_path: str) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    parts = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"Sheet: {sheet}\n" + "\n".join(rows))

    return "\n\n".join(parts) if parts else "[Empty spreadsheet]"


def _read_csv_rows(content: str, dialect: Any) -> list[str]:
    stream = io.StringIO(content)
    reader = csv.reader(stream, dialect) if dialect else csv.reader(stream)
    rows = []
    for row in reader:
        if any(c.strip() for c in row):
            rows.append(" | ".join(row))
    return rows


async def _extract_csv(file_path: str) -> str:
    with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
        content = f.read()

    try:
        dialect = csv.Sniffer().sniff(content[:2048])
    except csv.Error:
        dialect = None

    # The sniffer can guess a bogus delimiter on files with title/preamble rows
    # (common in real vendor quotes), which then makes csv.reader raise. Fall
    # back to the default comma dialect rather than failing the whole extraction.
    try:
        rows = _read_csv_rows(content, dialect)
    except csv.Error:
        rows = _read_csv_rows(content, None)

    return "\n".join(rows) if rows else "[Empty CSV]"
